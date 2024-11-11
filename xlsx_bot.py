import os
import psycopg2
from telegram import Update, InputFile
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters, CallbackContext
from openpyxl import load_workbook
from datetime import datetime, date
from pydantic import BaseModel
from typing import Optional, List
import warnings

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

TELEGRAM_TOKEN = ''
DATABASE_CONFIG = {
    'dbname': os.environ["POSTGRES_DATABASE"],
    'user': os.environ["POSTGRES_USER"],
    'password': os.environ["POSTGRES_PASSWORD"],
    'host': os.environ["POSTGRES_HOST"],
    'port': os.environ["POSTGRES_PORT"]
}


class Payment(BaseModel):
    account_number: Optional[str]
    operation_type: Optional[str]
    transaction_date: Optional[date]
    amount: Optional[float]
    payment_purpose: Optional[str]
    recipient_inn: Optional[str]
    recipient_name: Optional[str]
    counterparty_account: Optional[str]
    counterparty_inn: Optional[str]
    counterparty_name: Optional[str]
    counterparty_bank_bik: Optional[str]


def get_data_from_excel(file_path: str) -> List[Payment]:
    workbook = load_workbook(filename=file_path)
    sheet = workbook.worksheets[0]

    processed_data = []

    for row in sheet.iter_rows(min_row=11, max_row=sheet.max_row):
        if all(cell.value is None for cell in row):
            continue

        transaction_date_str = row[2].value
        transaction_date = None
        if transaction_date_str:
            try:
                transaction_date = datetime.strptime(transaction_date_str, "%d.%m.%Y").date()
            except ValueError:
                print(f"Warning: Could not parse date string '{transaction_date_str}'. Setting it to None.")

        processed_row = Payment(
            account_number=row[0].value,
            operation_type=row[1].value,
            transaction_date=transaction_date,
            amount=row[5].value,
            payment_purpose=row[8].value,
            recipient_inn=row[17].value,
            recipient_name=row[19].value,
            counterparty_account=row[22].value,
            counterparty_inn=row[23].value,
            counterparty_name=row[24].value,
            counterparty_bank_bik=row[25].value
        )
        processed_data.append(processed_row)

    return processed_data


def insert_into_postgres(conn, payments: List[Payment]):
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS payments (
            id SERIAL PRIMARY KEY,
            account_number VARCHAR(50),
            operation_type VARCHAR(50),
            transaction_date DATE,
            amount NUMERIC,
            payment_purpose TEXT,
            recipient_inn VARCHAR(20),
            recipient_name TEXT,
            counterparty_account VARCHAR(50),
            counterparty_inn VARCHAR(20),
            counterparty_name TEXT,
            counterparty_bank_bik VARCHAR(20),
            CONSTRAINT unique_payment UNIQUE (account_number, operation_type, transaction_date, amount, payment_purpose, recipient_inn, recipient_name, counterparty_account, counterparty_inn, counterparty_name, counterparty_bank_bik)
        );
    """)

    conn.commit()

    upsert_query = """
        INSERT INTO payments (
            account_number, operation_type, transaction_date, amount, payment_purpose, 
            recipient_inn, recipient_name, counterparty_account, counterparty_inn, 
            counterparty_name, counterparty_bank_bik
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (account_number, operation_type, transaction_date, amount, payment_purpose, 
                     recipient_inn, recipient_name, counterparty_account, counterparty_inn, 
                     counterparty_name, counterparty_bank_bik)
        DO NOTHING;
    """

    for payment in payments:
        data_tuple = (
            payment.account_number,
            payment.operation_type,
            payment.transaction_date,
            payment.amount,
            payment.payment_purpose,
            payment.recipient_inn,
            payment.recipient_name,
            payment.counterparty_account,
            payment.counterparty_inn,
            payment.counterparty_name,
            payment.counterparty_bank_bik
        )

        cur.execute(upsert_query, data_tuple)

    conn.commit()
    cur.close()


def handle_file(update: Update, context: CallbackContext):
    file = context.bot.get_file(update.message.document.file_id)
    file_path = f"{file.file_id}.xlsx"
    file.download(file_path)

    try:
        payments = get_data_from_excel(file_path)
        conn = psycopg2.connect(**DATABASE_CONFIG)
        insert_into_postgres(conn, payments)
        conn.close()
        update.message.reply_text("File processed and data uploaded to Grafana.")
    except Exception as e:
        update.message.reply_text(f"An error occurred: {e}")
    finally:
        os.remove(file_path)


def start(update: Update, context: CallbackContext):
    update.message.reply_text("Welcome! Send an .xlsx file to upload data to Grafana.")


def main():
    updater = Updater(TELEGRAM_TOKEN, use_context=True)
    dp = updater.dispatcher

    dp.add_handler(CommandHandler("start", start))
    dp.add_handler(
        MessageHandler(Filters.document.mime_type("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                       handle_file))

    updater.start_polling()
    updater.idle()


if __name__ == '__main__':
    main()
